"""
DIGEMID Alertas Scraper — scripts/scraper.py
Parte del skill digemid-alertas.

Uso standalone:
    python3 scraper.py                        # 1 página, motor heurístico
    ANTHROPIC_API_KEY=sk-... python3 scraper.py  # 1 página, Claude API

Uso como módulo (desde skill):
    exec(open('scripts/scraper.py').read())
    df = scrapear_alertas(max_paginas=2, analizar_acciones=True)
    exportar_excel(df, '/tmp/alertas.xlsx')
"""

# ── Auto-install de dependencias ───────────────────────────────────────────────
import subprocess, sys

def _ensure(pkg, import_name=None):
    try:
        __import__(import_name or pkg)
    except ImportError:
        subprocess.check_call([sys.executable, '-m', 'pip', 'install', pkg,
                               '--break-system-packages', '-q'])

_ensure('requests')
_ensure('bs4', 'bs4')
_ensure('pandas')
_ensure('openpyxl')
_ensure('pymupdf', 'fitz')

# ── Imports ────────────────────────────────────────────────────────────────────
import os, re, json, time
import requests
from bs4 import BeautifulSoup
import pandas as pd
from datetime import datetime
import fitz   # pymupdf

# ── Constantes ─────────────────────────────────────────────────────────────────
BASE_URL    = "https://www.digemid.minsa.gob.pe"
ALERTAS_URL = f"{BASE_URL}/alertas"
PAGE_URL    = f"{BASE_URL}/webDigemid/alertas/page/{{page}}/"
PDF_DELAY   = 15   # segundos entre descargas de PDF (rate limit del servidor)

# Sesión con headers de navegador real para evitar Cloudflare 403
_session = requests.Session()
_HEADERS = {
    "User-Agent":      "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Accept":          "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "es-PE,es;q=0.9,en-US;q=0.8",
    "Accept-Encoding": "gzip, deflate, br",
    "Referer":         "https://www.digemid.minsa.gob.pe/",
}
_session.get(BASE_URL, headers=_HEADERS, timeout=30)   # calentar cookies

# ── Prompt Claude API ──────────────────────────────────────────────────────────
_PROMPT = """\
Eres un experto en regulación farmacéutica peruana.
Analiza el texto de una Alerta DIGEMID y extrae las acciones requeridas.
Responde SOLO con JSON válido, sin texto adicional ni bloques markdown:

{
  "accion_principal": "frase corta, ej: RETIRO DEL MERCADO | NO COMERCIALIZAR | NOTIFICAR REACCIONES ADVERSAS",
  "urgencia": "INMEDIATA | PREVENTIVA | INFORMATIVA",
  "dirigido_a": ["destinatario 1", "destinatario 2"],
  "acciones_detalladas": ["acción concreta 1", "acción 2"],
  "resumen_accion": "1-2 oraciones resumiendo qué debe hacer el lector"
}

TEXTO:
{texto}
"""

# ══════════════════════════════════════════════════════════════════════════════
# MOTOR 1: CLAUDE API
# ══════════════════════════════════════════════════════════════════════════════
def _analizar_claude(texto: str) -> dict | None:
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        return None
    try:
        import anthropic
        client = anthropic.Anthropic(api_key=api_key)
        msg = client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=800,
            messages=[{"role": "user",
                        "content": _PROMPT.replace("{texto}", texto[:8000])}]
        )
        raw = msg.content[0].text.strip()
        raw = re.sub(r"^```json\s*|^```\s*|\s*```$", "", raw, flags=re.MULTILINE).strip()
        return json.loads(raw)
    except Exception as e:
        print(f"    [Claude API] {e}")
        return None


# ══════════════════════════════════════════════════════════════════════════════
# MOTOR 2: HEURÍSTICO (fallback sin API key)
# ══════════════════════════════════════════════════════════════════════════════
def _analizar_heuristico(texto: str) -> dict:
    """Extractor basado en patrones. Cubre los 3 tipos principales de DIGEMID."""
    up = texto.upper()

    # Acción principal
    if any(k in up for k in ["FALSIFICADO", "INCAUTADO", "FALSIFICACIÓN"]):
        accion, urgencia = "NO ADQUIRIR / NO COMERCIALIZAR", "INMEDIATA"
    elif any(k in up for k in ["RETIRO DEL MERCADO", "RECALL", "RETIRAR DEL MERCADO"]):
        accion, urgencia = "RETIRO DEL MERCADO", "INMEDIATA"
    elif any(k in up for k in ["SUSPENDER", "SUSPENSIÓN", "PROHIBIR"]):
        accion, urgencia = "SUSPENDER USO / DISTRIBUCIÓN", "INMEDIATA"
    elif any(k in up for k in ["RESULTADO CRÍTICO", "NO CONFORME", "SUBESTÁNDAR"]):
        accion, urgencia = "RETIRO POR CONTROL DE CALIDAD", "INMEDIATA"
    elif any(k in up for k in ["RIESGO", "REACCIÓN ADVERSA", "SIADM", "NOTIF"]):
        accion, urgencia = "NOTIFICAR / MEDIDAS PREVENTIVAS", "PREVENTIVA"
    else:
        accion, urgencia = "VER COMUNICADO OFICIAL", "INFORMATIVA"

    # Dirigido a: secciones "A los/las ..."
    dirigido = list(dict.fromkeys([
        d.strip() for d in
        re.findall(r"[Aa]\s+los?(?:as)?\s+([\w\s,áéíóúñÁÉÍÓÚÑ]+?):", texto)
    ]))[:4]

    # Acciones: viñetas •  o  -
    bullets = [
        l.strip().lstrip("•").lstrip("-").strip()
        for l in texto.split("\n")
        if l.strip().startswith(("•", "-")) and len(l.strip()) > 15
    ][:8]

    # Resumen
    resumen = ""
    for pat in [r"[Ss]e recomienda\s+([^.]+\.)",
                r"[Ss]e (solicita|requiere|exige|dispone)\s+([^.]+\.)",
                r"no (comprar|adquirir|utilizar|comercializar)\s+([^.]+\.)"]:
        m = re.search(pat, texto, re.IGNORECASE)
        if m:
            resumen = m.group(0).strip()
            break
    if not resumen and bullets:
        resumen = bullets[0][:220]

    return {"accion_principal": accion, "urgencia": urgencia,
            "dirigido_a": dirigido, "acciones_detalladas": bullets,
            "resumen_accion": resumen}


# ══════════════════════════════════════════════════════════════════════════════
# PDF: DESCARGA + EXTRACCIÓN DE TEXTO
# ══════════════════════════════════════════════════════════════════════════════
def _obtener_pdf_url(url_alerta: str) -> str | None:
    try:
        r = _session.get(url_alerta, headers=_HEADERS, timeout=20)
        r.raise_for_status()
        soup = BeautifulSoup(r.text, "html.parser")
        entry = soup.find("div", class_="entry-content")
        if not entry:
            return None
        link = entry.find("a", href=re.compile(r"\.pdf$", re.I))
        if not link:
            embed = entry.find("embed", src=re.compile(r"\.pdf$", re.I))
            if embed:
                link_href = embed.get("src", "")
                return link_href if link_href.startswith("http") else f"{BASE_URL}{link_href}"
            return None
        href = link["href"]
        return href if href.startswith("http") else f"{BASE_URL}{href}"
    except Exception:
        return None


def _descargar_pdf(pdf_url: str, reintentos: int = 3) -> bytes | None:
    headers = {**_HEADERS, "Accept": "application/pdf,*/*"}
    for intento in range(1, reintentos + 1):
        try:
            r = _session.get(pdf_url, headers=headers, timeout=30)
            if r.status_code == 200:
                return r.content
            elif r.status_code == 429:
                espera = PDF_DELAY * intento
                print(f"    [429 rate-limit] esperando {espera}s...", end=" ", flush=True)
                time.sleep(espera)
            else:
                return None
        except Exception:
            time.sleep(PDF_DELAY)
    return None


def _extraer_texto_pdf(pdf_bytes: bytes) -> str:
    try:
        doc = fitz.open(stream=pdf_bytes, filetype="pdf")
        return "\n".join(p.get_text() for p in doc).strip()
    except Exception:
        return ""


# ══════════════════════════════════════════════════════════════════════════════
# ENRIQUECIMIENTO DE ALERTA (PDF + IA)
# ══════════════════════════════════════════════════════════════════════════════
def _enriquecer(alerta: dict) -> dict:
    """Descarga PDF y extrae acciones. Modifica el dict in-place y lo retorna."""
    url = alerta.get("url")
    if not url:
        alerta.update({"pdf_url": None, "accion_principal": "Sin URL",
                       "urgencia": "INFORMATIVA", "resumen_accion": "Sin URL",
                       "acciones_detalladas": "", "dirigido_a": "", "motor_analisis": "—"})
        return alerta

    pdf_url = _obtener_pdf_url(url)
    alerta["pdf_url"] = pdf_url

    if not pdf_url:
        alerta.update({"accion_principal": "Sin PDF", "urgencia": "INFORMATIVA",
                       "resumen_accion": "PDF no encontrado.",
                       "acciones_detalladas": "", "dirigido_a": "", "motor_analisis": "—"})
        return alerta

    pdf_bytes = _descargar_pdf(pdf_url)
    if not pdf_bytes:
        alerta.update({"accion_principal": "Error descarga PDF", "urgencia": "INFORMATIVA",
                       "resumen_accion": "No se pudo descargar el PDF.",
                       "acciones_detalladas": "", "dirigido_a": "", "motor_analisis": "—"})
        return alerta

    texto = _extraer_texto_pdf(pdf_bytes)
    if len(texto) < 50:
        alerta.update({"accion_principal": "PDF escaneado (sin texto)", "urgencia": "INFORMATIVA",
                       "resumen_accion": "El PDF parece ser imagen escaneada.",
                       "acciones_detalladas": "", "dirigido_a": "", "motor_analisis": "—"})
        return alerta

    resultado = _analizar_claude(texto) or _analizar_heuristico(texto)
    motor = "Claude API" if os.environ.get("ANTHROPIC_API_KEY") else "Heurístico"

    alerta["accion_principal"]    = resultado.get("accion_principal", "")
    alerta["urgencia"]            = resultado.get("urgencia", "INFORMATIVA")
    alerta["resumen_accion"]      = resultado.get("resumen_accion", "")
    alerta["acciones_detalladas"] = " | ".join(resultado.get("acciones_detalladas", []))
    alerta["dirigido_a"]          = " | ".join(resultado.get("dirigido_a", []))
    alerta["motor_analisis"]      = motor
    return alerta


# ══════════════════════════════════════════════════════════════════════════════
# PARSEO DE PÁGINAS
# ══════════════════════════════════════════════════════════════════════════════
def _parsear_articulos(soup: BeautifulSoup) -> list[dict]:
    alertas = []
    for article in soup.find_all("article", class_=re.compile(r"\bpost\b")):
        titulo_tag = article.find("h2", class_="entry-title")
        if not titulo_tag:
            continue
        titulo = titulo_tag.get_text(strip=True)
        if "ALERTA DIGEMID" not in titulo.upper():
            continue
        link_tag = titulo_tag.find("a")
        link = link_tag["href"] if link_tag and link_tag.get("href") else None
        time_tag = article.find("time")
        fecha_pub = None
        if time_tag and time_tag.get("datetime"):
            try:
                fecha_pub = datetime.strptime(time_tag["datetime"], "%Y-%m-%d").date()
            except ValueError:
                pass
        excerpt = article.find("p", class_="post-excerpt")
        producto = excerpt.get_text(strip=True) if excerpt else None
        cat_tags = article.select("div.post-meta span.meta-cats a[rel='category tag']")
        cats = [c.get_text(strip=True) for c in cat_tags
                if c.get_text(strip=True) not in ("Alertas", "Alertas y Modificaciones")]
        alertas.append({
            "titulo": titulo, "producto": producto,
            "tipo_alerta": cats[0] if cats else "General",
            "fecha_publicacion": fecha_pub, "url": link,
            "fecha_captura": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        })
    return alertas


def _total_paginas(soup: BeautifulSoup) -> int:
    pag = soup.find("div", class_="pagination")
    if not pag:
        return 1
    nums = [int(a.get_text(strip=True))
            for a in pag.find_all("a", class_="page-numbers")
            if a.get_text(strip=True).isdigit()]
    return max(nums) if nums else 1


# ══════════════════════════════════════════════════════════════════════════════
# FUNCIÓN PRINCIPAL PÚBLICA
# ══════════════════════════════════════════════════════════════════════════════
def scrapear_alertas(max_paginas: int = 1,
                     delay_paginas: float = 1.5,
                     analizar_acciones: bool = True) -> pd.DataFrame:
    """
    Extrae alertas de DIGEMID con paginación y análisis opcional de acciones.

    Args:
        max_paginas       : Nº de páginas del listado (10 alertas/página).
                            None = todas (165 páginas, ~1,650 alertas).
        delay_paginas     : Segundos entre requests de páginas del listado.
        analizar_acciones : Si True, descarga PDFs y extrae acciones requeridas.

    Returns:
        pd.DataFrame con todas las alertas y (si analizar_acciones=True) las
        columnas: accion_principal, urgencia, dirigido_a, acciones_detalladas,
        resumen_accion, motor_analisis, pdf_url.
    """
    # Página 1
    resp = _session.get(ALERTAS_URL, headers=_HEADERS, timeout=30)
    resp.raise_for_status()
    soup = BeautifulSoup(resp.text, "html.parser")

    total = _total_paginas(soup)
    if max_paginas is not None:
        total = min(total, max_paginas)

    print(f"📋 Páginas a procesar: {total} ({total * 10} alertas aprox.)")
    alertas = _parsear_articulos(soup)
    print(f"  Página 1: {len(alertas)} alertas")

    for pag in range(2, total + 1):
        try:
            r = _session.get(PAGE_URL.format(page=pag), headers=_HEADERS, timeout=30)
            r.raise_for_status()
            nuevas = _parsear_articulos(BeautifulSoup(r.text, "html.parser"))
            alertas.extend(nuevas)
            print(f"  Página {pag}/{total}: {len(nuevas)} alertas")
        except Exception as e:
            print(f"  ERROR página {pag}: {e}")
        time.sleep(delay_paginas)

    if analizar_acciones:
        motor = "Claude API" if os.environ.get("ANTHROPIC_API_KEY") else "Heurístico"
        print(f"\n🔍 Analizando acciones [{motor}] para {len(alertas)} alertas...")
        print(f"   (delay entre PDFs: {PDF_DELAY}s — servidor DIGEMID tiene rate limit)\n")
        for i, a in enumerate(alertas, 1):
            print(f"  [{i:>3}/{len(alertas)}] {a['titulo']}...", end=" ", flush=True)
            _enriquecer(a)
            urgencia = a.get("urgencia", "—")
            emoji = {"INMEDIATA": "🔴", "PREVENTIVA": "🟡", "INFORMATIVA": "🔵"}.get(urgencia, "⚪")
            print(f"{emoji} [{urgencia:11}] {a.get('accion_principal', '?')}")
            time.sleep(PDF_DELAY)

    df = pd.DataFrame(alertas)
    print(f"\n✅ Total alertas: {len(df)}")
    if analizar_acciones and "urgencia" in df.columns:
        print(df["urgencia"].value_counts().to_string())
    return df


# ══════════════════════════════════════════════════════════════════════════════
# EXPORTAR A EXCEL
# ══════════════════════════════════════════════════════════════════════════════
def exportar_excel(df: pd.DataFrame, ruta: str):
    """Genera Excel formateado con colores por urgencia e hipervínculos."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    COLS = [
        ("Título",               "titulo",               40),
        ("Producto",             "producto",             28),
        ("Tipo de Alerta",       "tipo_alerta",          20),
        ("Fecha Publicación",    "fecha_publicacion",    15),
        ("⚡ Acción Principal",  "accion_principal",     30),
        ("Urgencia",             "urgencia",             13),
        ("Dirigido a",           "dirigido_a",           35),
        ("Acciones Requeridas",  "acciones_detalladas",  65),
        ("Resumen IA",           "resumen_accion",       55),
        ("Motor Análisis",       "motor_analisis",       14),
        ("URL Alerta",           "url",                  50),
        ("URL PDF",              "pdf_url",              50),
        ("Fecha Captura",        "fecha_captura",        16),
    ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Alertas DIGEMID"

    thin   = Side(style="thin", color="BBBBBB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    h_fill = PatternFill("solid", start_color="1F4E79")
    h_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    d_font = Font(name="Arial", size=9)

    for ci, (label, _, width) in enumerate(COLS, 1):
        c = ws.cell(row=1, column=ci, value=label)
        c.font = h_font; c.fill = h_fill; c.border = border
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[1].height = 24

    URGENCIA_BG    = {"INMEDIATA": "FCE4D6", "PREVENTIVA": "FFF2CC", "INFORMATIVA": "EBF3FB"}
    URGENCIA_COLOR = {"INMEDIATA": "C00000", "PREVENTIVA": "ED7D31", "INFORMATIVA": "2E75B6"}

    for ri, row in enumerate(df.itertuples(index=False), 2):
        urgencia  = getattr(row, "urgencia", None) or "INFORMATIVA"
        row_fill  = PatternFill("solid", start_color=URGENCIA_BG.get(urgencia, "F2F2F2"))
        for ci, (_, field, _) in enumerate(COLS, 1):
            val = getattr(row, field, None)
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = d_font; c.fill = row_fill; c.border = border
            c.alignment = Alignment(wrap_text=True, vertical="center")
            if field in ("url", "pdf_url") and val:
                c.hyperlink = str(val)
                c.font = Font(name="Arial", size=9, color="0563C1", underline="single")
            if field == "urgencia" and val:
                c.font = Font(name="Arial", size=9, bold=True,
                              color=URGENCIA_COLOR.get(urgencia, "000000"))
                c.alignment = Alignment(horizontal="center", vertical="center")
            if field == "accion_principal":
                c.font = Font(name="Arial", size=9, bold=True)
        ws.row_dimensions[ri].height = 52

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}{len(df)+1}"

    # Hoja resumen
    ws2 = wb.create_sheet("Resumen")
    ws2["A1"] = "Reporte Alertas DIGEMID"
    ws2["A1"].font = Font(bold=True, size=13, name="Arial", color="1F4E79")
    ws2["A3"] = "Fecha captura:"; ws2["B3"] = datetime.now().strftime("%d/%m/%Y %H:%M")
    ws2["A4"] = "Total alertas:"; ws2["B4"] = len(df)
    ws2["A6"] = "Por urgencia:"; ws2["A6"].font = Font(bold=True, name="Arial")
    if "urgencia" in df.columns:
        for i, (k, v) in enumerate(df["urgencia"].value_counts().items(), 7):
            ws2[f"A{i}"] = k; ws2[f"B{i}"] = v
    ws2["D6"] = "Acciones principales:"; ws2["D6"].font = Font(bold=True, name="Arial")
    if "accion_principal" in df.columns:
        for i, (k, v) in enumerate(df["accion_principal"].value_counts().head(8).items(), 7):
            ws2[f"D{i}"] = k; ws2[f"E{i}"] = v
    for col in ["A","B","D","E"]:
        ws2.column_dimensions[col].width = 40

    wb.save(ruta)
    print(f"💾 Excel guardado: {ruta}")


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    motor = "Claude API" if os.environ.get("ANTHROPIC_API_KEY") else "Heurístico"
    print(f"Motor: {motor}\n")

    df = scrapear_alertas(max_paginas=1, analizar_acciones=True)

    fecha = datetime.now().strftime("%Y%m%d_%H%M")
    ruta  = f"/tmp/alertas_digemid_{fecha}.xlsx"
    exportar_excel(df, ruta)
